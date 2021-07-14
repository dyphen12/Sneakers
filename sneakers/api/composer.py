"""

Prisma Inc. 2021

composer.py

Status: Checked

Note: For unit testing purposes.

Made by Alexis W.

"""
import pandas as pd
import json
from openpyxl import load_workbook
from sneakers.api import utils
from sneakers.api import processing
from sneakers.api import core
from sneakers.api import uploaders


def chunks(data, n):

    m = int(len(data)/n)

    return [data[x:x+m] for x in range(0, len(data), m)]


# Composer is intended to work only on an Excel Worksheet object environment.
class Composer:
    def __init__(self, title, samplesize=0):
        self.exist = False
        self.samplesize = samplesize
        self.folder = 'workout'
        self.title = title
        self.doc_file = '{ftit}.xlsx'.format(ftit=title)
        self.full_path = 'workout/{ftit}.xlsx'.format(ftit=title)
        self.doc_id = 'NaN'
        self.online = False
        self.json_name = 'workout/{}.json'.format(title)
        
        
        try:
            wb = load_workbook(filename=self.full_path)
            self.exist = True
        except FileNotFoundError as e:
            print('Composer workbook not found.')
            print('Creating workbook...')
            self.create_workbook()
            print('Workbook Created!')
            print('Check {}'.format(self.full_path))
            self.exist = True
            
            
        if self.exist == True:
            try:
                with open(self.json_name) as jsonFile:
                    jsonObject = json.load(jsonFile)
                    jsonFile.close()
            
                if jsonObject['composer']['doc_id'] == 'NaN':
                    self.online = False
                    
                else:
                    self.doc_id = jsonObject['composer']['doc_id']
                    self.online = True
                    
            except FileNotFoundError:
                
                print('JSON does not exist')
            


    def create_workbook(self):
        # Creates a new empty workbook with no sneaker images, just data
        dataset = utils.load_shoes_dataset()
        sample = dataset.iloc[:self.samplesize]

        sample['pic'] = ''
        writer = pd.ExcelWriter(self.full_path, engine='xlsxwriter', options={'strings_to_urls': False})
        sample.to_excel(writer)
        writer.close()
        
        aid = 'NaN'
        
        self.doc_id = aid
        
        json_comp = {"composer":{"doc_id":"NaN"}}
            
        with open('workout/{}.json'.format(self.title), 'w', encoding='utf-8') as f:
                
            json.dump(json_comp, f, ensure_ascii=False, indent=4)
            
        print('Workbook saved to ', self.full_path)



    def load_wb(self):
        wb = load_workbook(filename=self.full_path)
        return wb

    def write_wb(self, addr, local=False):
        msg = 'Writing from {ffrom} to {fto}'.format(ffrom=addr[0], fto=addr[1])
        print(msg)
        df = utils.load_shoes_dataset()
        sample = df.iloc[:self.samplesize]
        wb = self.load_wb()
        ws = wb.active
        processes = processing.img_pre_multiprocess(sample, ws, self.full_path)

        selection = []
        writings = []

        for i in range(addr[0], addr[1]+1):
            cellname = 'S{fnum}'.format(fnum=i)
            selection.append(cellname)

        for cell in selection:
            for i in range(0, len(processes)):
                if cell == processes[i][0]:
                    writings.append(processes[i])

        if local == True:
            processing.img_processor(writings, local=True)
        else:
            processing.img_processor(process_list=writings)

        print('XLSX processed successfully')

        return True

    def write_wb_xl(self, addr, iny_size=10, local=False):
        msg = 'Writing from {ffrom} to {fto}'.format(ffrom=addr[0], fto=addr[1])
        print(msg)
        df = utils.load_shoes_dataset()
        sample = df.iloc[:self.samplesize]
        wb = self.load_wb()
        ws = wb.active
        processes = processing.img_pre_multiprocess(sample, ws, self.full_path)

        selection = []
        writings = []

        for i in range(addr[0], addr[1]+1):
            cellname = 'S{fnum}'.format(fnum=i)
            selection.append(cellname)

        for cell in selection:
            for i in range(0, len(processes)):
                if cell == processes[i][0]:
                    writings.append(processes[i])

        if local == True:
            processing.img_processor_inj(writings, iny_size)
        else:
            processing.img_processor_inj(process_list=writings, size=iny_size, local=False)

        print('XLSX processed successfully')

        return True

    # CONNECTS TO CORE
    def update_prices(self):

        wb = self.load_wb()
        ws = wb.active

        # iterate through excel and display data
        # iterate through excel and display data
        for i in range(0, self.samplesize):

            price_cell = 'C{fnum}'.format(fnum=i+2)
            id_cell = 'E{fnum}'.format(fnum=i + 2)

            id = ws[id_cell]
            sneaker = core.get_shoe_by_id(id)
            try:
                new_price = sneaker[0]['estimatedMarketValue']
            except IndexError:
                continue
            ws[price_cell] = new_price

        wb.save(self.full_path)
        try:
            wb.save(self.full_path)
        except PermissionError:
            print('It seems like the xlsx is being used by another program, please close it first and try again.')

        print('Prices Updated Succesufully')
        return True
        
    def upload_file(self):
        
        # Just for the fist time
        
        
        if self.online == False:
            
        
            aid = uploaders.upload_folder(self.full_path, self.doc_file)
            self.doc_id = aid
            
            json_comp = {
                            "composer": {
                                "doc_id": aid
                            }
                        }
            
            with open(self.json_name, 'w', encoding='utf-8') as f:
                
                json.dump(json_comp, f, ensure_ascii=False, indent=4)
                
                
            
            self.online = True
            
            
        else:
            
            print('This file is already online')
            
        
        
        return True
        
    
    def sync_file(self):
        
        if self.online == True:
        
            uploaders.sync_by_id(self.doc_id, self.full_path, self.doc_file)
        
        else:
            
            print('Sync failed, this workbook is not in the cloud.')
            print('You can upload it with composer.upload_file() .')
        
        return True
        
        
    #def check_json():
        
        















