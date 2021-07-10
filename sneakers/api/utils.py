"""

Made by Alexis Wong.
Prisma Inc.

"""

import pandas as pd
import os
import requests
import progressbar
import time
import gc
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as pyImage

from sneakers.api import processing


# Status: Check
def load_shoes_dataset():
    shoes = pd.read_csv('sneakers/datasets/sneakers-100000-extended.csv', index_col=0, low_memory=False)
    return shoes


# Status: Check
def flush_sheets():

    arr = os.listdir('sheets')

    for file in arr:
        os.remove('sheets/{fname}'.format(fname=file))

    return True


# Status: Check
def flush_outputs():

    arr = os.listdir('outputs')

    for file in arr:
        os.remove('outputs/{fname}'.format(fname=file))

    return True


# Status: Check
def flush_img():

    arr = os.listdir('img')

    for file in arr:
        os.remove('img/{fname}'.format(fname=file))

    return True


# Status: Check
def build_xlsx(df, ver):

    title = "sneakers-{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    df.to_excel(path)

    wb = load_workbook(filename=path)

    ws = wb.active

    progress_lenc = len(df.index)

    progsc = progressbar

    for i in progsc.progressbar(range(0, progress_lenc)):

        TestUrl = df['image'][i]

        cellName = 'S{fnum}'.format(fnum=i + 2)

        try:

            im = Image.open(requests.get(TestUrl, stream=True).raw)

            im_100 = im.resize((78, 100))

            loc = "img/Sneaker{}.png".format(i + 2)

            im_100.save(loc, format="png")

            img = Image.open(loc)

            xImg = pyImage(img)

            ws[cellName] = ''

            ws.add_image(xImg, cellName)

        except:

            continue

    wb.save(path)
    print('Images downloaded succesfully...')
    print('XLSX builded...')
    print('Check /sheets folder for {ffile}'.format(ffile = path))

    return wb


# Status: Check
def build_large_xlsx(df, ver):

    title = "sneakers-{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    df['pic'] = ''

    writer = pd.ExcelWriter(path, engine='xlsxwriter', options={'strings_to_urls': False})
    df.to_excel(writer)
    writer.close()

    wb = load_workbook(filename=path)

    progress_lenc = len(df.index)

    progsc = progressbar

    for i in progsc.progressbar(range(0, progress_lenc)):

        ws = wb.active

        TestUrl = df['image'][i]

        j = i + 2

        cellName = 'S{fnum}'.format(fnum=j)

        try:

            im = Image.open(requests.get(TestUrl, stream=True).raw)

            # time.sleep(3)

            im_100 = im.resize((78, 100))

            loc = "img/Sneaker{}.png".format(j)

            im_100.save(loc, format="png")

            img = Image.open(loc)

            xImg = pyImage(img)

            ws[cellName] = ''

            ws.add_image(xImg, cellName)

            wb.save(path)

            del img
            del xImg
            del cellName
            del TestUrl
            del loc
            del im_100
            del im

            gc.collect()

        except requests.exceptions.MissingSchema:

            print('MissingSchema Error on iteration number {fit}'.format(fit=j))

            continue

        except requests.exceptions.ConnectionError:

            print('Connection Error on iteration number {fit}'.format(fit=j))

            continue

    print('Images downloaded succesfully...')
    print('XLSX large builded...')
    print('Check /sheets folder for {ffile}'.format(ffile=path))

    return True


# These functions are Jupyter Compilant Only
def get_images(df):
    for index, row in df.iterrows():
        TestUrl = row['image']
        # im = Image.open(requests.get(TestUrl, stream=True).raw)
        row['image'] = TestUrl

    return df


def get_images_to(df):
    for index, row in df.iterrows():
        TestUrl = row['image']
        im = Image.open(requests.get(TestUrl, stream=True))
        print(im)

    return df


# FUNCTION DEVELOPMENT

def build_big_xlsx(df, ver):

    title = "sneakers-{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    df['pic'] = ''

    writer = pd.ExcelWriter(path, engine='xlsxwriter', options={'strings_to_urls': False})
    df.to_excel(writer)
    writer.close()

    wb = load_workbook(filename=path)

    ws = wb.active

    processing.processing_xlsx(df, ws)

    wb.save(path)

    print('Images downloaded succesfully...')
    print('XLSX large builded...')
    print('Check /sheets folder for {ffile}'.format(ffile=path))

    return True