"""

Prisma Inc. 2021

injector.py

Status: Checked

Note: Injector class for image adding.

Made by Alexis W.

"""
import progressbar
import time
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as pyImage


def chunks(data, n):

    m = int(len(data)/n)

    return [data[x:x+m] for x in range(0, len(data), m)]


class cylinder:
    def __init__(self, images, size):
        self.images = images
        self.size = size
        self.space = len(images)

    def injection(self):

        inj_q = int(self.space/self.size)

        print('{finjq} inyectors loaded'.format(finjq=inj_q))

        batch = chunks(self.images, inj_q)

        for k in range(0, len(batch)):
            print('Batch {kf} loaded into chamber. Size: {fs}'.format(kf=k, fs=len(batch[k])))
            self.chamber(batch[k])

        return 'yay'

    @staticmethod
    def chamber(images):

        time.sleep(1)

        path = images[0][0][2]

        wb = load_workbook(filename=path, keep_links=False)

        # Progress Bar Object
        progsq = progressbar

        for i in progsq.progressbar(range(0, len(images))):

            loc = images[i][0][0]

            try:
                cellname = images[i][0][1]
            except IndexError:
                continue

            try:

                imgd = Image.open(loc)

                xImg = pyImage(imgd)

                ws = wb.active

                ws[cellname] = ''

                ws.add_image(xImg, cellname)

            # This Exception is only raised when running local=True
            except FileNotFoundError:

                continue

            wb.save(path)
            wb.close()

        return True