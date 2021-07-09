"""

Made by Alexis Wong.
Prisma Inc.

"""

from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as pyImage
from PIL import Image
import progressbar
import requests


# Status: Failed // Unicode Error, presumably can't decode the symbol ö from Möebius.
def upload_as_csv(df, ver):

    title = "sneakers-{fver}.csv".format(fver=ver)

    path = 'outputs/{ftit}'.format(ftit=title)

    df.to_csv(path)

    gauth = GoogleAuth()
    drive = GoogleDrive(gauth)

    uploaded = drive.CreateFile({'title': title})
    uploaded.SetContentFile(path)
    uploaded.Upload()
    print('Uploaded file with ID {}'.format(uploaded.get('id')))
    print('Version number {}'.format(ver))

    return True


# Status: Failed // Unicode Error, presumably can't decode the symbol ö from Möebius.

def upload_as_xlsx(df, ver):

    title = "sneakers-{fver}.xlsx".format(fver=ver)

    path = 'sheets/{ftit}'.format(ftit=title)

    df.to_excel(path)

    wb = load_workbook(filename=path)

    ws = wb.active

    progress_lenc = len(df.index)

    progsc = progressbar

    for i in progsc.progressbar(range(0, progress_lenc)):

        TestUrl = df['image'][i]

        cellName = 'G{fnum}'.format(fnum=i + 3)

        # print(cellName)

        # img = Image.open(BytesIO(response.content))

        try:


            im = Image.open(requests.get(TestUrl, stream=True).raw)

            im_100 = im.resize((78, 100))

            loc = "img/Sneaker{}.png".format(i + 2)

            im_100.save(loc, format="png")

            img = Image.open(loc)

            xImg = pyImage(img)
            # print(img)

            ws[cellName] = ''

            ws.add_image(xImg, cellName)

        except:

            continue

    print('Images downloaded succesfully...')
    print('XLSX builded...')
    print('Check your drive account')

    wb.save(path)
    gauth = GoogleAuth()
    drive = GoogleDrive(gauth)

    uploaded = drive.CreateFile({'title': title})
    uploaded.SetContentFile(path)
    uploaded.Upload()
    print('Uploaded file with ID {}'.format(uploaded.get('id')))
    print('Version number {}'.format(ver))

    return None